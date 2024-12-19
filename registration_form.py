import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import os
import re

def validate_date(date_text):
    """
    Comprehensive date validation with separate error checks:
    1. Correct format (DD/MM/YYYY)
    2. Year range check
    3. Month and day validity
    
    Returns:
    - True if valid
    - Specific error message if invalid
    """
    try:
        # Check format first
        if not re.match(r'\d{2}/\d{2}/\d{4}', date_text):
            return "Invalid date format. Use DD/MM/YYYY"
        
        # Split the date
        day, month, year = map(int, date_text.split('/'))
        
        # Year validation
        if year < 1950 or year > 2024:
            return "Year must be between 1950 and 2024"
        
        # Month validation
        if month < 1 or month > 12:
            return "Invalid month. Must be between 1 and 12"
        
        # Days in each month (accounting for leap years)
        days_in_month = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 
                         31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        
        # Day validation
        if day < 1 or day > days_in_month[month - 1]:
            return f"Invalid day for {['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'][month - 1]}. Must be between 1 and {days_in_month[month - 1]}"
        
        return True
    except ValueError:
        return "Invalid date. Please check your input"

def submit_form():
    name = entry_name.get()
    father_name = entry_father_name.get()
    mother_name = entry_mother_name.get()
    permanent_address = entry_permanent_address.get()
    current_address = entry_current_address.get()
    dob = entry_dob.get()
    dept = entry_dept.get()
    marks_10th = entry_10th_marks.get()
    marks_12th = entry_12th_marks.get()

    # Basic validation for empty fields
    required_fields = [
        (name, "Name"),
        (father_name, "Father's Name"),
        (mother_name, "Mother's Name"),
        (permanent_address, "Permanent Address"),
        (current_address, "Current Address"),
        (dob, "Date of Birth"),
        (dept, "Department"),
        (marks_10th, "10th Marks"),
        (marks_12th, "12th Marks")
    ]

    for field, field_name in required_fields:
        if not field:
            messagebox.showerror("Error", f"{field_name} field required")
            return

    # Validate marks
    try:
        marks_10th = int(marks_10th)
        marks_12th = int(marks_12th)
        if not (0 <= marks_10th <= 1000):
            messagebox.showerror("Error", "10th Marks must be between 0 and 1000")
            return
        if not (0 <= marks_12th <= 1000):
            messagebox.showerror("Error", "12th Marks must be between 0 and 1000")
            return
    except ValueError:
        messagebox.showerror("Error", "Marks must be an integer")
        return

    # Validate date of birth
    date_validation = validate_date(dob)
    if date_validation is not True:
        messagebox.showerror("Invalid Date of Birth", date_validation)
        return

    # Define the file path
    filepath = "registration_data.xlsx"

    # Check if the file exists
    if not os.path.exists(filepath):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Registration Data"
        sheet.append(["Name", "Father's Name", "Mother's Name", "Permanent Address", "Current Address", "Date of Birth", "Department", "10th Marks", "12th Marks"])
    else:
        workbook = load_workbook(filepath)
        sheet = workbook.active

    # Append the data
    sheet.append([name, father_name, mother_name, permanent_address, current_address, dob, dept, marks_10th, marks_12th])

    # Save the workbook
    workbook.save(filepath)
    messagebox.showinfo("Success", "Registration Successful")
    clear_form()

def clear_form():
    for entry in entries:
        entry.delete(0, tk.END)

# Create the main window
root = tk.Tk()
root.title("Student Registration Form")

# Create labels and entry widgets for the form fields
labels_text = [
    "Name:", "Father's Name:", "Mother's Name:", 
    "Permanent Address:", "Current Address:", 
    "Date of Birth (DD/MM/YYYY):", "Department:", 
    "10th Marks:", "12th Marks:"
]

entries = []
for i, label_text in enumerate(labels_text):
    label = tk.Label(root, text=label_text)
    label.grid(row=i, column=0, padx=10, pady=5, sticky='w')
    
    entry = tk.Entry(root, width=30)
    entry.grid(row=i, column=1, padx=10, pady=5)
    entries.append(entry)

# Unpack entries for easier reference
(entry_name, entry_father_name, entry_mother_name, 
 entry_permanent_address, entry_current_address, 
 entry_dob, entry_dept, entry_10th_marks, entry_12th_marks) = entries

# Button frame for better spacing
button_frame = tk.Frame(root)
button_frame.grid(row=len(labels_text), column=0, columnspan=2, pady=10)

# Create styled buttons with custom appearance
submit_button = tk.Button(
    button_frame, 
    text="Submit", 
    command=submit_form,
    bg='lightblue',
    fg='black',
    relief=tk.RAISED,
    borderwidth=2,
    padx=10,
    pady=5
)
submit_button.pack(side=tk.RIGHT, padx=10)

clear_button = tk.Button(
    button_frame, 
    text="Clear", 
    command=clear_form,
    bg='lightcoral',
    fg='black',
    relief=tk.RAISED,
    borderwidth=2,
    padx=10,
    pady=5
)
clear_button.pack(side=tk.LEFT, padx=10)

root.mainloop()
